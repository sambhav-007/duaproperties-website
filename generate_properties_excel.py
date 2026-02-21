"""
Dua Properties - Excel Sheet Generator
Generates a comprehensive Excel workbook from properties.json
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Colours ──────────────────────────────────────────────────────────────────
HEADER_BG      = "1F3864"   # dark navy
HEADER_FONT    = "FFFFFF"
SUB_HEADER_BG  = "2E75B6"   # mid blue
SUB_HEADER_FG  = "FFFFFF"
ALT_ROW_BG     = "EBF3FB"   # light blue tint
WHITE          = "FFFFFF"
ACCENT_GOLD    = "C9A84C"
SECTION_BG     = "D6E4F7"
BORDER_COLOR   = "BFBFBF"

def thin_border():
    s = Side(border_style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def header_border():
    s = Side(border_style="medium", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=HEADER_BG, fg=HEADER_FONT, size=11, bold=True, wrap=True):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)
    cell.border    = header_border()

def style_data(cell, row_idx, wrap=False, bold=False, color="000000"):
    bg = ALT_ROW_BG if row_idx % 2 == 0 else WHITE
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Calibri", size=10, bold=bold, color=color)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border    = thin_border()

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ─────────────────────────────────────────────────────────────────────────────

def load_data(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe(val, default="—"):
    """Return a printable string; collapse lists with newlines."""
    if val is None:
        return default
    if isinstance(val, list):
        return "\n".join(str(v) for v in val) if val else default
    return str(val) if str(val).strip() else default

def join_list(lst, sep=", "):
    if not lst:
        return "—"
    return sep.join(str(x) for x in lst)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – PROPERTIES OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════

OVERVIEW_COLS = [
    ("ID",              12),
    ("Property Name",   42),
    ("Developer",       28),
    ("Type",            20),
    ("Sub-Type",        24),
    ("Configuration",   30),
    ("Location",        40),
    ("Price",           26),
    ("Price Value (₹)", 20),
    ("Price Unit",      16),
    ("Status",          14),
    ("Possession",      24),
    ("BHK Type",        18),
    ("Bedrooms",        12),
    ("Bathrooms",       12),
    ("Area (Sq.Ft)",    16),
    ("Plot Size (Sq.Yd)",18),
    ("Furnishing",      18),
    ("RERA / Approval", 30),
    ("Tags",            40),
    ("Video URL",       40),
]

def build_overview(ws, data):
    # Title row
    ws.merge_cells("A1:U1")
    title = ws["A1"]
    title.value = "DUA PROPERTIES — Complete Property Listings"
    title.font  = Font(name="Calibri", bold=True, size=16, color=HEADER_FONT)
    title.fill  = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Sub-title row
    ws.merge_cells("A2:U2")
    sub = ws["A2"]
    sub.value = f"Total Properties Listed: {len(data)}   |   Generated for Client Reference"
    sub.font  = Font(name="Calibri", bold=False, size=11, color=SUB_HEADER_FG)
    sub.fill  = PatternFill("solid", fgColor=SUB_HEADER_BG)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Column headers
    for col_idx, (header, width) in enumerate(OVERVIEW_COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        style_header(cell, bg=SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 28

    # Data rows
    for row_idx, prop in enumerate(data, start=1):
        r = row_idx + 3
        ws.row_dimensions[r].height = 18

        def w(col, val, wrap=False, bold=False, color="000000"):
            c = ws.cell(row=r, column=col, value=val)
            style_data(c, row_idx, wrap=wrap, bold=bold, color=color)

        plot_size = prop.get("plot_size") or (
            f"{prop.get('min_plot_size_sqyd','?')}–{prop.get('max_plot_size_sqyd','?')}"
            if prop.get("min_plot_size_sqyd") else "—"
        )

        w(1,  safe(prop.get("id")))
        w(2,  safe(prop.get("name")),  bold=True, color="1F3864")
        w(3,  safe(prop.get("developer")))
        w(4,  safe(prop.get("type")))
        w(5,  safe(prop.get("subType")))
        w(6,  safe(prop.get("configuration") or prop.get("bhkType")))
        w(7,  safe(prop.get("location")), wrap=True)
        w(8,  safe(prop.get("priceDisplay") or prop.get("price")), bold=True, color="1F5C00")
        w(9,  safe(prop.get("priceValue")))
        w(10, safe(prop.get("priceUnit")))
        w(11, safe(prop.get("status")))
        w(12, safe(prop.get("possessionStatus") or prop.get("possession")))
        w(13, safe(prop.get("bhkType")))
        w(14, safe(prop.get("bedrooms")))
        w(15, safe(prop.get("bathrooms")))
        w(16, safe(prop.get("area_sqft")))
        w(17, plot_size)
        w(18, safe(prop.get("furnishingStatus")))
        w(19, safe(prop.get("rera_id")))
        w(20, join_list(prop.get("tags", [])), wrap=True)
        w(21, safe(prop.get("video_url")))

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(OVERVIEW_COLS))}3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – HIGHLIGHTS & DESCRIPTION
# ══════════════════════════════════════════════════════════════════════════════

def build_highlights(ws, data):
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Property Highlights & Descriptions"
    title.font  = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT)
    title.fill  = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [("ID", 8), ("Property Name", 40), ("Description", 80), ("Key Highlights", 60)]
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        style_header(cell, bg=SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 24

    for row_idx, prop in enumerate(data, 1):
        r = row_idx + 2
        highlights = prop.get("highlights", [])
        desc = safe(prop.get("description"))
        hl_text = "\n".join(f"• {h}" for h in highlights) if highlights else "—"

        for col, val, wrap in [
            (1, safe(prop.get("id")), False),
            (2, safe(prop.get("name")), True),
            (3, desc, True),
            (4, hl_text, True),
        ]:
            cell = ws.cell(row=r, column=col, value=val)
            style_data(cell, row_idx, wrap=wrap)

        # Row height proportional to content
        lines = max(desc.count("\n") + 1, len(highlights), 1)
        ws.row_dimensions[r].height = min(max(lines * 14, 30), 200)

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – AMENITIES
# ══════════════════════════════════════════════════════════════════════════════

def build_amenities(ws, data):
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "Property Amenities"
    title.font  = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT)
    title.fill  = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [("ID", 8), ("Property Name", 42), ("Amenities", 90)]
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        style_header(cell, bg=SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 24

    for row_idx, prop in enumerate(data, 1):
        r = row_idx + 2
        amenities = prop.get("amenities", [])
        am_text = "\n".join(f"✓  {a}" for a in amenities) if amenities else "—"

        for col, val, wrap in [
            (1, safe(prop.get("id")), False),
            (2, safe(prop.get("name")), True),
            (3, am_text, True),
        ]:
            cell = ws.cell(row=r, column=col, value=val)
            style_data(cell, row_idx, wrap=wrap)

        ws.row_dimensions[r].height = min(max(len(amenities) * 14, 30), 300)

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – PAYMENT PLANS
# ══════════════════════════════════════════════════════════════════════════════

def build_payment(ws, data):
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "Payment Plans"
    title.font  = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT)
    title.fill  = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [("ID",8),("Property Name",42),("Basic Price",28),
               ("Plot Sizes Available",30),("Payment Stages",60),("Notes",60)]
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        style_header(cell, bg=SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 24

    for row_idx, prop in enumerate(data, 1):
        r = row_idx + 2
        pp = prop.get("payment_plan", {})
        basic = pp.get("basic_price") or pp.get("basic_price_type") or "—"
        plot_sizes = join_list(pp.get("plot_sizes", []))
        stages = pp.get("stages", [])
        stages_text = "\n".join(
            f"{s.get('name','')}: {s.get('amount', s.get('details',''))}"
            for s in stages
        ) if stages else "—"
        notes_text = "\n".join(f"• {n}" for n in prop.get("notes", [])) or "—"

        for col, val, wrap in [
            (1, safe(prop.get("id")), False),
            (2, safe(prop.get("name")), True),
            (3, basic, True),
            (4, plot_sizes, True),
            (5, stages_text, True),
            (6, notes_text, True),
        ]:
            cell = ws.cell(row=r, column=col, value=val)
            style_data(cell, row_idx, wrap=wrap)

        lines = max(len(stages), notes_text.count("\n") + 1, 2)
        ws.row_dimensions[r].height = min(max(lines * 14, 30), 250)

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – OTHER CHARGES
# ══════════════════════════════════════════════════════════════════════════════

def build_charges(ws, data):
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Other Charges & Additional Costs"
    title.font  = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT)
    title.fill  = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [("ID",8),("Property Name",42),("Charge Name",36),("Amount / Details",40)]
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        style_header(cell, bg=SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 24

    current_row = 3
    for prop in data:
        charges = prop.get("other_charges", [])
        if not charges:
            charges = [{"name": "—", "details": "No additional charges listed"}]

        prop_id   = safe(prop.get("id"))
        prop_name = safe(prop.get("name"))

        for i, charge in enumerate(charges):
            amount = charge.get("amount") or charge.get("details") or "—"
            for col, val, wrap, bold, color in [
                (1, prop_id if i == 0 else "",   False, False, "000000"),
                (2, prop_name if i == 0 else "", True,  i==0,  "1F3864" if i==0 else "000000"),
                (3, safe(charge.get("name")),    True,  False, "000000"),
                (4, amount,                       True,  False, "000000"),
            ]:
                cell = ws.cell(row=current_row, column=col, value=val)
                row_parity = (current_row - 3) % 2
                style_data(cell, row_parity, wrap=wrap, bold=bold, color=color)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – NEARBY LOCATIONS
# ══════════════════════════════════════════════════════════════════════════════

def build_nearby(ws, data):
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Nearby Landmarks & Connectivity"
    title.font  = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT)
    title.fill  = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [("ID",8),("Property Name",42),("Landmark / Place",40),("Distance",20)]
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        style_header(cell, bg=SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 24

    current_row = 3
    wrote_any = False
    for prop in data:
        nearby = prop.get("nearby", [])
        if not nearby:
            continue
        wrote_any = True
        prop_id   = safe(prop.get("id"))
        prop_name = safe(prop.get("name"))

        for i, place in enumerate(nearby):
            for col, val, wrap, bold, color in [
                (1, prop_id if i == 0 else "",   False, False, "000000"),
                (2, prop_name if i == 0 else "", True,  i==0,  "1F3864" if i==0 else "000000"),
                (3, safe(place.get("name")),     True,  False, "000000"),
                (4, safe(place.get("distance")), False, False, "000000"),
            ]:
                cell = ws.cell(row=current_row, column=col, value=val)
                row_parity = (current_row - 3) % 2
                style_data(cell, row_parity, wrap=wrap, bold=bold, color=color)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    if not wrote_any:
        ws.cell(row=3, column=1, value="No nearby landmark data available for current listings.")

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 – SUMMARY STATS
# ══════════════════════════════════════════════════════════════════════════════

def build_summary(ws, data):
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 36

    def heading(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        style_header(cell, bg=HEADER_BG, size=12)
        ws.row_dimensions[row].height = 26

    def kv(row, key, val, highlight=False):
        k = ws.cell(row=row, column=1, value=key)
        v = ws.cell(row=row, column=2, value=val)
        bg = ACCENT_GOLD if highlight else (ALT_ROW_BG if row % 2 == 0 else WHITE)
        for c in (k, v):
            c.fill   = PatternFill("solid", fgColor=bg)
            c.font   = Font(name="Calibri", size=11, bold=highlight)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 20

    row = 1
    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "DUA PROPERTIES — Portfolio Summary"
    title.font  = Font(name="Calibri", bold=True, size=16, color=HEADER_FONT)
    title.fill  = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── General Stats ──
    row = 3
    heading(row, "General Statistics")
    row += 1
    kv(row, "Total Properties", len(data), highlight=True)
    row += 1

    type_counts = {}
    for p in data:
        t = p.get("type", "Unknown")
        type_counts[t] = type_counts.get(t, 0) + 1
    for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        kv(row, f"  → {t}", cnt)
        row += 1

    # ── Possession ──
    row += 1
    heading(row, "Possession Status Breakdown")
    row += 1
    poss_counts = {}
    for p in data:
        ps = p.get("possessionStatus") or p.get("possession") or "Not specified"
        poss_counts[ps] = poss_counts.get(ps, 0) + 1
    for ps, cnt in sorted(poss_counts.items(), key=lambda x: -x[1]):
        kv(row, f"  → {ps}", cnt)
        row += 1

    # ── Locations ──
    row += 1
    heading(row, "Locations")
    row += 1
    for p in data:
        kv(row, safe(p.get("name"))[:45], safe(p.get("location")))
        row += 1

    # ── Price Range ──
    row += 1
    heading(row, "Price Overview")
    row += 1
    for p in data:
        price_disp = safe(p.get("priceDisplay") or p.get("price"))
        kv(row, safe(p.get("name"))[:45], price_disp)
        row += 1

    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path  = os.path.join(script_dir, "src", "data", "properties.json")
    out_path   = os.path.join(script_dir, "Dua_Properties_Listings.xlsx")

    print(f"Loading data from: {json_path}")
    data = load_data(json_path)
    print(f"Loaded {len(data)} properties.")

    wb = Workbook()

    # Rename default sheet
    ws1 = wb.active
    ws1.title = "📋 Overview"

    ws2 = wb.create_sheet("📝 Highlights")
    ws3 = wb.create_sheet("🏊 Amenities")
    ws4 = wb.create_sheet("💳 Payment Plans")
    ws5 = wb.create_sheet("💰 Other Charges")
    ws6 = wb.create_sheet("📍 Nearby")
    ws7 = wb.create_sheet("📊 Summary")

    print("Building sheets...")
    build_overview(ws1, data)
    build_highlights(ws2, data)
    build_amenities(ws3, data)
    build_payment(ws4, data)
    build_charges(ws5, data)
    build_nearby(ws6, data)
    build_summary(ws7, data)

    # Set tab colours
    tab_colors = ["1F3864","2E75B6","217346","7030A0","C55A11","375623","843C0C"]
    for ws, color in zip([ws1,ws2,ws3,ws4,ws5,ws6,ws7], tab_colors):
        ws.sheet_properties.tabColor = color

    wb.save(out_path)
    print(f"\n✅ Excel file saved to:\n   {out_path}")

if __name__ == "__main__":
    main()
